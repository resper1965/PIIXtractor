import os
import re
import logging
import time
from dataclasses import dataclass
from typing import Dict, Any

from validate_docbr import CPF, CNPJ
from docx import Document
from openpyxl import load_workbook
from PyPDF2 import PdfReader
from tqdm import tqdm

from extractor.openai_classifier import classify_text


REGEX_PATTERNS = {
    "email": re.compile(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b"),
    "cpf": re.compile(r"\b\d{3}\.?\d{3}\.?\d{3}-?\d{2}\b"),
    "cnpj": re.compile(r"\b\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}\b"),
    "telefone": re.compile(r"\b(?:\(?\d{2}\)?\s?)?(?:9\d{4}|\d{4})-?\d{4}\b"),
    "data_nasc": re.compile(r"\b(0?[1-9]|[12][0-9]|3[01])[/-](0?[1-9]|1[0-2])[/-](\d{2}|\d{4})\b"),
    "cep": re.compile(r"\b\d{5}-?\d{3}\b"),
    "nome": re.compile(r"\b([A-ZÀ-Ú][a-zà-ú]+(?:\s[A-ZÀ-Ú][a-zà-ú]+)+)\b"),
}


def validate_cpf(value: str) -> bool:
    """Validate CPF by stripping non-digit characters."""
    return CPF().validate(re.sub(r"\D", "", value))


def validate_cnpj(value: str) -> bool:
    """Validate CNPJ by stripping non-digit characters."""
    return CNPJ().validate(re.sub(r"\D", "", value))


def extract_text(path: str) -> str:
    """Read and return the content of supported files."""
    try:
        if path.endswith(".txt"):
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        if path.endswith(".pdf"):
            reader = PdfReader(path)
            return "\n".join(p.extract_text() or "" for p in reader.pages)
        if path.endswith(".docx"):
            doc = Document(path)
            return "\n".join(p.text for p in doc.paragraphs)
        if path.endswith(".xlsx"):
            wb = load_workbook(path)
            data = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    data.append(" ".join(str(c) for c in row if c))
            return "\n".join(data)
    except Exception as exc:
        logging.warning("[ERRO] %s: %s", path, exc)
    return ""


def extract_regex(text: str) -> Dict[str, Any]:
    """Extract data using regex patterns."""
    result: Dict[str, Any] = {}
    for key, pattern in REGEX_PATTERNS.items():
        matches = pattern.findall(text)
        if key == "cpf":
            matches = [v for v in matches if validate_cpf(v)]
        elif key == "cnpj":
            matches = [v for v in matches if validate_cnpj(v)]
        if matches:
            result[key] = sorted(set(matches))
    return result


@dataclass
class Progress:
    total: int
    bar: tqdm

    @classmethod
    def create(cls, total: int) -> "Progress":
        bar = tqdm(total=total, desc="Processando arquivos", unit="arquivo")
        return cls(total=total, bar=bar)

    def update(self) -> None:
        self.bar.update(1)

    def close(self) -> None:
        self.bar.close()


def process_directory(directory: str) -> Dict[str, Any]:
    """Process all supported files inside the given directory."""
    all_files = [os.path.join(root, f)
                 for root, _, files in os.walk(directory)
                 for f in files
                 if f.lower().endswith((".txt", ".pdf", ".docx", ".xlsx"))]

    progress = Progress.create(len(all_files))
    results: Dict[str, Any] = {}

    for path in all_files:
        logging.info("Processando: %s", path)
        text = extract_text(path)
        if not text.strip():
            progress.update()
            continue
        if len(text) > 50000:
            logging.warning("[SKIP] %s excede 50k caracteres. Ignorando.", path)
            progress.update()
            continue
        regex_result = extract_regex(text)
        gpt_result = ""
        if not regex_result:
            gpt_result = classify_text(text)
        results[path] = {
            "regex": regex_result,
            "gpt": gpt_result,
        }
        progress.update()
        time.sleep(2)

    progress.close()
    return results
