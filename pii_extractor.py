import os
import re
import zipfile
import logging
import shutil
import time
import csv
import json
import sqlite3
from dataclasses import dataclass
from typing import Dict, Any

from dotenv import load_dotenv
from validate_docbr import CPF, CNPJ
from docx import Document
from openpyxl import load_workbook
from PyPDF2 import PdfReader
from extractor.openai_classifier import classify_text
from tqdm import tqdm

# --- Initial Configuration ---
load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)

REGEX_PATTERNS = {
    "email": re.compile(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b"),
    "cpf": re.compile(r"\b\d{3}\.?\d{3}\.?\d{3}-?\d{2}\b"),
    "cnpj": re.compile(r"\b\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}\b"),
    "telefone": re.compile(r"\b(?:\(?\d{2}\)?\s?)?(?:9\d{4}|\d{4})-?\d{4}\b"),
    "data_nasc": re.compile(r"\b(0?[1-9]|[12][0-9]|3[01])[/-](0?[1-9]|1[0-2])[/-](\d{2}|\d{4})\b"),
    "cep": re.compile(r"\b\d{5}-?\d{3}\b"),
    "nome": re.compile(r"\b([A-ZÀ-Ú][a-zà-ú]+(?:\s[A-ZÀ-Ú][a-zà-ú]+)+)\b")
}

def validar_cpf(valor: str) -> bool:
    """Valida CPF removendo caracteres não numéricos."""
    return CPF().validate(re.sub(r"\D", "", valor))


def validar_cnpj(valor: str) -> bool:
    """Valida CNPJ removendo caracteres não numéricos."""
    return CNPJ().validate(re.sub(r"\D", "", valor))


def extrair_texto_arquivo(caminho: str) -> str:
    """Lê e retorna o conteúdo de arquivos suportados."""
    try:
        if caminho.endswith(".txt"):
            with open(caminho, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        if caminho.endswith(".pdf"):
            reader = PdfReader(caminho)
            return "\n".join([p.extract_text() or "" for p in reader.pages])
        if caminho.endswith(".docx"):
            doc = Document(caminho)
            return "\n".join([p.text for p in doc.paragraphs])
        if caminho.endswith(".xlsx"):
            wb = load_workbook(caminho)
            dados = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    dados.append(" ".join([str(c) for c in row if c]))
            return "\n".join(dados)
    except Exception as e:
        logging.warning(f"[ERRO] {caminho}: {e}")
    return ""


def extrair_regex(texto: str) -> Dict[str, Any]:
    """Extrai dados utilizando regex."""
    resultado = {}
    for tipo, pattern in REGEX_PATTERNS.items():
        encontrados = pattern.findall(texto)
        if tipo == "cpf":
            encontrados = [v for v in encontrados if validar_cpf(v)]
        elif tipo == "cnpj":
            encontrados = [v for v in encontrados if validar_cnpj(v)]
        if encontrados:
            resultado[tipo] = sorted(set(encontrados))
    return resultado




@dataclass
class Progresso:
    total: int
    barra: tqdm

    @classmethod
    def criar(cls, total: int) -> 'Progresso':
        barra = tqdm(total=total, desc="Processando arquivos", unit="arquivo")
        return cls(total=total, barra=barra)

    def atualizar(self):
        self.barra.update(1)

    def finalizar(self):
        self.barra.close()


def processar_diretorio(diretorio: str) -> Dict[str, Any]:
    """Processa todos os arquivos do diretório informado."""
    todos_arquivos = [os.path.join(r, f)
                      for r, _, files in os.walk(diretorio)
                      for f in files
                      if f.lower().endswith((".txt", ".pdf", ".docx", ".xlsx"))]
    progresso = Progresso.criar(len(todos_arquivos))

    resultados = {}
    for caminho in todos_arquivos:
        logging.info(f"Processando: {caminho}")
        texto = extrair_texto_arquivo(caminho)
        if not texto.strip():
            progresso.atualizar()
            continue
        if len(texto) > 50000:
            logging.warning(f"[SKIP] {caminho} excede 50k caracteres. Ignorando.")
            progresso.atualizar()
            continue
        regex_result = extrair_regex(texto)
        gpt_result = ""
        if not regex_result:
            gpt_result = classify_text(texto)
        resultados[caminho] = {
            "regex": regex_result,
            "gpt": gpt_result
        }
        progresso.atualizar()
        time.sleep(2)

    progresso.finalizar()
    return resultados


def exportar_resultados_csv(resultados: Dict[str, Any], caminho_csv: str = "resultados.csv") -> None:
    """Exporta resultados para arquivo CSV."""
    with open(caminho_csv, mode="w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["arquivo", "origem", "tipo", "valor"])
        for arquivo, dados in resultados.items():
            for origem in ["regex", "gpt"]:
                info = dados.get(origem)
                if isinstance(info, dict):
                    for tipo, valores in info.items():
                        if isinstance(valores, list):
                            for valor in valores:
                                writer.writerow([arquivo, origem, tipo, valor])
                        else:
                            writer.writerow([arquivo, origem, tipo, valores])
                else:
                    writer.writerow([arquivo, origem, "conteudo", info])


def exportar_resultados_json(resultados: Dict[str, Any], caminho_json: str = "resultados.json") -> None:
    """Exporta resultados para arquivo JSON."""
    with open(caminho_json, "w", encoding="utf-8") as f:
        json.dump(resultados, f, indent=2, ensure_ascii=False)


def exportar_resultados_sqlite(resultados: Dict[str, Any], db_path: str = "resultados.db") -> None:
    """Exporta resultados para banco SQLite."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS dados (
            arquivo TEXT,
            origem TEXT,
            tipo TEXT,
            valor TEXT
        )
        """
    )
    for arquivo, dados in resultados.items():
        for origem in ["regex", "gpt"]:
            info = dados.get(origem)
            if isinstance(info, dict):
                for tipo, valores in info.items():
                    if isinstance(valores, list):
                        for valor in valores:
                            cursor.execute(
                                "INSERT INTO dados VALUES (?, ?, ?, ?)",
                                (arquivo, origem, tipo, valor)
                            )
                    else:
                        cursor.execute(
                            "INSERT INTO dados VALUES (?, ?, ?, ?)",
                            (arquivo, origem, tipo, valores)
                        )
            else:
                cursor.execute(
                    "INSERT INTO dados VALUES (?, ?, ?, ?)",
                    (arquivo, origem, "conteudo", info)
                )
    conn.commit()
    conn.close()


def extrair_e_processar_zip(zip_path: str) -> Dict[str, Any]:
    """Extrai arquivos de um ZIP e processa seu conteúdo."""
    tmpdir = "/tmp/piiextractor_tmp"
    os.makedirs(tmpdir, exist_ok=True)

    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(tmpdir)
            logging.info(f"ZIP extraído em: {tmpdir}")
    except Exception as e:
        logging.error(f"[ERRO ao extrair ZIP] {e}")
        return {}

    resultados = processar_diretorio(tmpdir)

    try:
        shutil.rmtree(tmpdir)
        logging.info(f"Diretório temporário removido: {tmpdir}")
    except Exception as e:
        logging.warning(f"Falha ao remover tmpdir: {e}")

    return resultados


if __name__ == "__main__":
    zip_file = os.getenv("ZIP_INPUT", "columbiati.zip")
    resultados = extrair_e_processar_zip(zip_file)

    for arq, info in resultados.items():
        print(f"\n\U0001F4C4 {arq}")
        print("Regex:", info["regex"])
        print("GPT:", info["gpt"])

    exportar_resultados_csv(resultados, caminho_csv="resultados.csv")
    exportar_resultados_json(resultados, caminho_json="resultados.json")
    exportar_resultados_sqlite(resultados, db_path="resultados.db")
    logging.info("Exportação concluída: CSV, JSON e SQLite.")
